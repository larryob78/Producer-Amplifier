"""XLSX export from BreakdownOutput using openpyxl.

Produces a styled workbook with:
- Summary sheet: project info and global flags
- Scenes sheet: one row per scene with wrapped cells
- Hidden Costs sheet: severity, flag, affected scenes, mitigations
- Questions sheet: grouped by department
"""

from __future__ import annotations

from pathlib import Path

from sba.output.schema import BreakdownOutput


def export_xlsx(breakdown: BreakdownOutput, output_path: Path) -> Path:
    """Export breakdown as a styled XLSX workbook.

    Returns the path to the written file.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: " "pip install openpyxl"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12)
    label_font = Font(bold=True)

    ws_summary["A1"] = "Project Summary"
    ws_summary["A1"].font = header_font

    ps = breakdown.project_summary
    summary_rows = [
        ("Title", ps.project_title),
        ("Date Analyzed", ps.date_analyzed),
        ("Scope", ps.analysis_scope),
        ("Est. Pages", str(ps.script_pages_estimate or "—")),
        ("Total Scenes", str(ps.total_scene_count)),
    ]
    for i, (label, val) in enumerate(summary_rows, 3):
        ws_summary.cell(row=i, column=1, value=label).font = label_font
        ws_summary.cell(row=i, column=2, value=val)

    gf = breakdown.global_flags
    ws_summary.cell(row=9, column=1, value="Global Flags").font = header_font
    flag_rows = [
        ("VFX Heaviness", gf.overall_vfx_heaviness),
        ("Virtual Production Fit", gf.likely_virtual_production_fit),
        ("Top Risk Themes", ", ".join(gf.top_risk_themes) if gf.top_risk_themes else "—"),
    ]
    for i, (label, val) in enumerate(flag_rows, 11):
        ws_summary.cell(row=i, column=1, value=label).font = label_font
        ws_summary.cell(row=i, column=2, value=val)

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 50

    # --- Scenes sheet ---
    ws_scenes = wb.create_sheet("Scenes")
    scene_headers = [
        "Scene ID",
        "Slugline",
        "Int/Ext",
        "Day/Night",
        "Pages (1/8)",
        "Location",
        "Characters",
        "Summary",
        "VFX Categories",
        "Triggers",
        "Flags",
        "Shots Min",
        "Shots Likely",
        "Shots Max",
        "Invisible VFX",
        "Cost Risk",
        "Schedule Risk",
        "Risk Reasons",
        "Suggested Capture",
        "Notes for Producer",
        "Uncertainties",
    ]

    header_fill = PatternFill(start_color="2D2B27", end_color="2D2B27", fill_type="solid")
    header_text = Font(bold=True, color="E8E4DE")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, h in enumerate(scene_headers, 1):
        cell = ws_scenes.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill

    for row_idx, scene in enumerate(breakdown.scenes, 2):
        flags = scene.production_flags
        active = [
            k
            for k, v in [
                ("stunts", flags.stunts),
                ("creatures", flags.creatures),
                ("vehicles", flags.vehicles),
                ("crowds", flags.crowds),
                ("water", flags.water),
                ("fire_smoke", flags.fire_smoke),
                ("destruction", flags.destruction),
                ("weather", flags.weather),
            ]
            if v
        ]
        est = scene.vfx_shot_count_estimate
        values = [
            scene.scene_id,
            scene.slugline,
            scene.int_ext,
            scene.day_night,
            scene.page_count_eighths,
            scene.location_type,
            ", ".join(scene.characters),
            scene.scene_summary,
            ", ".join(c.value for c in scene.vfx_categories),
            ", ".join(scene.vfx_triggers),
            ", ".join(active),
            est.min,
            est.likely,
            est.max,
            scene.invisible_vfx_likelihood,
            scene.cost_risk_score,
            scene.schedule_risk_score,
            "\n".join(scene.risk_reasons),
            "\n".join(scene.suggested_capture),
            "\n".join(
                scene.notes_for_producer
                if isinstance(scene.notes_for_producer, list)
                else [scene.notes_for_producer]
            ),
            "\n".join(scene.uncertainties),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_scenes.cell(row=row_idx, column=col, value=val)
            cell.alignment = wrap

    # Auto-width approximation
    for col_idx in range(1, len(scene_headers) + 1):
        ws_scenes.column_dimensions[ws_scenes.cell(row=1, column=col_idx).column_letter].width = (
            min(40, max(12, len(scene_headers[col_idx - 1]) + 4))
        )

    # --- Hidden Costs sheet ---
    ws_costs = wb.create_sheet("Hidden Costs")
    cost_headers = ["Severity", "Flag", "Affected Scenes", "Why It Matters", "Mitigations"]
    for col, h in enumerate(cost_headers, 1):
        cell = ws_costs.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill

    for row_idx, item in enumerate(breakdown.hidden_cost_radar, 2):
        values = [
            item.severity,
            item.flag,
            ", ".join(item.where),
            item.why_it_matters,
            "\n".join(item.mitigation_ideas),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_costs.cell(row=row_idx, column=col, value=val)
            cell.alignment = wrap

    for col_idx, h in enumerate(cost_headers, 1):
        ws_costs.column_dimensions[ws_costs.cell(row=1, column=col_idx).column_letter].width = min(
            50, max(15, len(h) + 4)
        )

    # --- Questions sheet ---
    ws_qs = wb.create_sheet("Questions")
    q_headers = ["Department", "Question"]
    for col, h in enumerate(q_headers, 1):
        cell = ws_qs.cell(row=1, column=col, value=h)
        cell.font = header_text
        cell.fill = header_fill

    kq = breakdown.key_questions_for_team
    departments = [
        ("Producer", kq.for_producer),
        ("VFX Supervisor", kq.for_vfx_supervisor),
        ("DP / Camera", kq.for_dp_camera),
        ("Locations / Art Dept", kq.for_locations_art_dept),
    ]
    row = 2
    for dept, questions in departments:
        for q in questions:
            ws_qs.cell(row=row, column=1, value=dept)
            ws_qs.cell(row=row, column=2, value=q).alignment = wrap
            row += 1

    ws_qs.column_dimensions["A"].width = 24
    ws_qs.column_dimensions["B"].width = 80

    # Save
    wb.save(output_path)
    return output_path
