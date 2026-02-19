"""Excel budget reader — reads the master budget workbook via openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

# Module-level budget cache
_workbook: openpyxl.Workbook | None = None
_budget_path: Path | None = None

# Standard Hollywood account code → category mapping
ACCOUNT_NAMES = {
    "1100": "Story, Rights & Continuity",
    "1200": "Producers Unit",
    "1300": "Director",
    "1400": "Cast / Principal Talent",
    "1500": "Bits & Stunts",
    "1800": "ATL Travel & Living",
    "2000": "Production Staff",
    "2100": "Extra Talent / Background",
    "2200": "Set Design & Art Department",
    "2300": "Set Construction",
    "2400": "Set Striking",
    "2500": "Set Operations",
    "2600": "Special Effects (Mechanical)",
    "2700": "Set Dressing",
    "2800": "Property (Props)",
    "2900": "Wardrobe",
    "3000": "Picture Vehicles",
    "3100": "Makeup & Hairdressing",
    "3200": "Lighting",
    "3300": "Camera",
    "3400": "Production Sound",
    "3500": "Transportation",
    "3600": "Locations",
    "3700": "Production Film & Lab",
    "4300": "Visual Effects (VFX)",
    "4400": "Titles & Graphics",
    "4600": "Music",
    "5100": "Editing & Post Production",
    "5200": "Post Sound (ADR, Foley, Mix)",
    "6500": "Publicity",
    "6700": "Insurance",
    "6800": "General & Administrative",
    "7600": "Amortization",
    "9000": "Contingency",
}


def load_budget(path: str | Path) -> None:
    """Load the master budget Excel workbook into memory.

    Args:
        path: Path to the .xlsx master budget file.
    """
    global _workbook, _budget_path
    _budget_path = Path(path)
    _workbook = openpyxl.load_workbook(_budget_path, data_only=True)


def get_workbook() -> openpyxl.Workbook:
    """Return the loaded workbook or raise if not loaded."""
    if _workbook is None:
        raise RuntimeError("Budget not loaded. Call load_budget() first.")
    return _workbook


def read_account(account_code: str) -> dict[str, Any]:
    """Read a single account from the Top Sheet.

    Scans column A for the account code, then reads columns B-F.

    Returns:
        Dict with account_code, category, budget, actual, variance, variance_pct.
    """
    wb = get_workbook()
    ws = wb["Top Sheet"]

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and str(cell_val).strip() == str(account_code).strip():
            category = ws.cell(row=row, column=2).value or ACCOUNT_NAMES.get(
                account_code, "Unknown"
            )
            budget = ws.cell(row=row, column=3).value or 0
            actual = ws.cell(row=row, column=4).value or 0
            variance = ws.cell(row=row, column=5).value
            variance_pct = ws.cell(row=row, column=6).value

            # If variance wasn't calculated (formula not resolved), compute it
            if variance is None:
                variance = (budget or 0) - (actual or 0)
            if variance_pct is None and budget:
                variance_pct = variance / budget

            return {
                "account_code": account_code,
                "category": category,
                "budget": budget,
                "actual": actual,
                "variance": variance,
                "variance_pct": round(variance_pct * 100, 1) if variance_pct else 0,
                "status": "over" if (variance or 0) < 0 else "under",
            }

    return {
        "account_code": account_code,
        "error": f"Account {account_code} not found in Top Sheet",
    }


def get_budget_summary() -> dict[str, Any]:
    """Get a high-level budget summary for the system prompt context.

    Returns:
        Dict with total_budget, total_actual, total_variance, and section totals.
    """
    wb = get_workbook()
    ws = wb["Top Sheet"]

    sections = {}
    _current_section = None
    grand_budget = 0
    grand_actual = 0

    for row in range(1, ws.max_row + 1):
        _col_a = ws.cell(row=row, column=1).value  # noqa: F841
        col_b = ws.cell(row=row, column=2).value
        col_c = ws.cell(row=row, column=3).value
        col_d = ws.cell(row=row, column=4).value

        # Detect section headers
        if col_b and "TOTAL" in str(col_b).upper():
            if col_c and col_d:
                section_name = str(col_b).replace("TOTAL ", "").strip()
                sections[section_name] = {
                    "budget": col_c,
                    "actual": col_d,
                    "variance": (col_c or 0) - (col_d or 0),
                }

        # Detect grand total
        if col_b and "GRAND TOTAL" in str(col_b).upper():
            grand_budget = col_c or 0
            grand_actual = col_d or 0

    # Find over-budget accounts
    over_budget = []
    for row in range(1, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        if code and str(code).strip() in ACCOUNT_NAMES:
            budget = ws.cell(row=row, column=3).value or 0
            actual = ws.cell(row=row, column=4).value or 0
            if actual > budget and budget > 0:
                over_budget.append(
                    {
                        "account": str(code).strip(),
                        "category": ACCOUNT_NAMES.get(str(code).strip(), ""),
                        "budget": budget,
                        "actual": actual,
                        "overage": actual - budget,
                        "overage_pct": round((actual - budget) / budget * 100, 1),
                    }
                )

    # Fallback: sum sections if GRAND TOTAL row not found
    if grand_budget == 0 and sections:
        grand_budget = sum(float(s.get("budget", 0) or 0) for s in sections.values())
        grand_actual = sum(float(s.get("actual", 0) or 0) for s in sections.values())

    # Fallback: sum all account rows when GRAND TOTAL not found
    if grand_budget == 0:
        for row in range(1, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            if code and str(code).strip() in ACCOUNT_NAMES:
                try:
                    grand_budget += float(ws.cell(row=row, column=3).value or 0)
                    grand_actual += float(ws.cell(row=row, column=4).value or 0)
                except (ValueError, TypeError):
                    pass

    return {
        "total_budget": grand_budget,
        "total_actual": grand_actual,
        "total_variance": grand_budget - grand_actual,
        "sections": sections,
        "over_budget_accounts": sorted(over_budget, key=lambda x: x["overage"], reverse=True),
    }


def get_vfx_detail() -> list[dict[str, Any]]:
    """Read the VFX Detail sheet — per-scene VFX costs.

    Returns:
        List of dicts with scene, description, shots, and cost breakdown by category.
    """
    wb = get_workbook()
    if "VFX Detail" not in wb.sheetnames:
        return []

    ws = wb["VFX Detail"]
    scenes = []

    for row in range(4, ws.max_row + 1):
        scene_num = ws.cell(row=row, column=1).value
        if not scene_num or str(scene_num).strip() == "":
            continue

        desc = ws.cell(row=row, column=2).value
        if desc and "TOTAL" in str(desc).upper():
            break

        scenes.append(
            {
                "scene": str(scene_num).strip(),
                "description": desc,
                "shots": ws.cell(row=row, column=3).value or 0,
                "env_extension": ws.cell(row=row, column=4).value or 0,
                "creature_fx": ws.cell(row=row, column=5).value or 0,
                "wire_removal": ws.cell(row=row, column=6).value or 0,
                "particle_fx": ws.cell(row=row, column=7).value or 0,
                "cg_composite": ws.cell(row=row, column=8).value or 0,
                "total": ws.cell(row=row, column=9).value or 0,
            }
        )

    return scenes
