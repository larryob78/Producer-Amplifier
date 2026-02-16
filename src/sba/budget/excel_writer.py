"""Excel budget writer — writes changes back to the master budget with audit trail."""

from __future__ import annotations

import datetime
from typing import Any

import openpyxl

from sba.budget.excel_reader import ACCOUNT_NAMES, _budget_path, get_workbook


def update_account(
    account_code: str,
    field: str,
    value: Any,
    reason: str,
) -> dict[str, Any]:
    """Update a budget field and log the change.

    Args:
        account_code: The account code to update.
        field: Which field — 'budget', 'actual', or 'notes'.
        value: The new value.
        reason: Why the change was made (audit trail).

    Returns:
        Dict with old_value, new_value, and audit entry.
    """
    wb = get_workbook()
    ws = wb["Top Sheet"]

    field_col = {"budget": 3, "actual": 4, "notes": None}
    col = field_col.get(field)
    if col is None:
        return {"error": f"Unsupported field: {field}. Use 'budget' or 'actual'."}

    # Find the row
    target_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and str(cell_val).strip() == str(account_code).strip():
            target_row = row
            break

    if target_row is None:
        return {"error": f"Account {account_code} not found in Top Sheet"}

    old_value = ws.cell(row=target_row, column=col).value
    ws.cell(row=target_row, column=col, value=value)

    # Log to audit trail sheet
    _log_audit(wb, account_code, field, old_value, value, reason)

    # Save
    if _budget_path:
        wb.save(_budget_path)

    return {
        "account_code": account_code,
        "field": field,
        "old_value": old_value,
        "new_value": value,
        "reason": reason,
        "timestamp": datetime.datetime.now().isoformat(),
        "status": "updated",
    }


def _log_audit(
    wb: openpyxl.Workbook,
    account_code: str,
    field: str,
    old_value: Any,
    new_value: Any,
    reason: str,
) -> None:
    """Append an entry to the Audit Log sheet (creates it if missing)."""
    sheet_name = "Audit Log"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Timestamp", "Account", "Category", "Field", "Old Value", "New Value", "Reason", "User"])
    else:
        ws = wb[sheet_name]

    ws.append([
        datetime.datetime.now().isoformat(),
        account_code,
        ACCOUNT_NAMES.get(account_code, ""),
        field,
        old_value,
        new_value,
        reason,
        "Producer Copilot",
    ])
